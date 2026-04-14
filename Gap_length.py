import cv2
import numpy as np
from scipy import signal
from scipy.ndimage import median_filter
import matplotlib.pyplot as plt

class ImprovedGapDetector:
    def __init__(self, calibration_factor=None):

        self.calibration_factor = calibration_factor
        self.mm_per_pixel = (1.0 / calibration_factor) if calibration_factor else None

    def resize_image_to_target(self, image, target_mp=12.0):
        height, width = image.shape[:2]
        current_mp = (height * width) / 1e6

        if current_mp <= target_mp:
            return image

        scale = np.sqrt((target_mp * 1e6) / (height * width))
        new_width = max(1, int(width * scale))
        new_height = max(1, int(height * scale))

        resized = cv2.resize(image, (new_width, new_height), interpolation=cv2.INTER_AREA)

        print(f"Downscaled image from {width}x{height} ({current_mp:.2f} MP) "
              f"to {new_width}x{new_height} ({(new_width * new_height)/1e6:.2f} MP)")
        return resized

    def preprocess_for_vertical_gap(self, image):
        # Convert to grayscale
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

        # Apply median filter to reduce noise while preserving edges
        denoised = cv2.medianBlur(gray, ksize=3)

        # Enhance contrast using CLAHE
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(denoised)

        # Apply bilateral filter to smooth while keeping edges sharp
        bilateral = cv2.bilateralFilter(enhanced, 9, 75, 75)

        return gray, enhanced, bilateral

    def detect_vertical_edge(self, image):
        # Sobel X (vertical edges)
        sobelx = cv2.Sobel(image, cv2.CV_64F, 1, 0, ksize=3)
        sobelx_abs = np.abs(sobelx)

        # Scharr operator (more sensitive to edges)
        scharrx = cv2.Scharr(image, cv2.CV_64F, 1, 0)
        scharrx_abs = np.abs(scharrx)

        # Combine both
        combined_edges = (sobelx_abs + scharrx_abs) / 2

        # Normalize to 0-255
        edge_normalized = cv2.normalize(combined_edges, None, 0, 255, cv2.NORM_MINMAX, dtype=cv2.CV_8U)

        return edge_normalized

    def detect_vertical_edge_fast(self, image):
        """Fast vertical edge approx using Canny + Sobel for quick mode."""
        canny = cv2.Canny(image, 50, 150, apertureSize=3, L2gradient=True)
        sobelx = cv2.Sobel(image, cv2.CV_64F, 1, 0, ksize=3)
        sobelx_abs = np.abs(sobelx)
        sobel_norm = cv2.normalize(sobelx_abs, None, 0, 255, cv2.NORM_MINMAX, dtype=cv2.CV_8U)
        combined = cv2.addWeighted(canny, 0.6, sobel_norm, 0.4, 0)
        return combined

    def find_gap_center_line(self, edge_image, method='projection'):
        """Find the vertical center line of the gap"""
        height, width = edge_image.shape

        if method == 'projection':
            # Vertical projection - sum along vertical axis
            vertical_projection = np.sum(edge_image, axis=0)

            # Apply smoothing to projection (original accurate method)
            window_size = 11
            smoothed_projection = signal.savgol_filter(vertical_projection, window_size, 3)

            # Find peaks (strong vertical edges)
            from scipy.signal import find_peaks
            peaks, properties = find_peaks(smoothed_projection,
                                          prominence=np.max(smoothed_projection) * 0.2,
                                          distance=20)

            if len(peaks) == 0:
                # Fallback: find maximum
                gap_center_x = np.argmax(smoothed_projection)
            else:
                # Get the strongest peak
                strongest_peak_idx = np.argmax(properties['prominences'])
                gap_center_x = peaks[strongest_peak_idx]

            return gap_center_x, vertical_projection, smoothed_projection

        elif method == 'hough':
            # Use Hough Line Transform to detect vertical lines
            lines = cv2.HoughLines(edge_image, 1, np.pi/180, threshold=100)

            if lines is not None:
                # Filter for near-vertical lines
                vertical_lines = []
                for line in lines:
                    rho, theta = line[0]
                    # Check if line is near vertical (theta close to 0 or pi)
                    if abs(theta) < 0.1 or abs(theta - np.pi) < 0.1:
                        vertical_lines.append((rho, theta))

                if vertical_lines:
                    # Get the most central vertical line
                    rhos = [abs(rho) for rho, _ in vertical_lines]
                    center_line_idx = np.argmin([abs(rho - edge_image.shape[1]//2)
                                                for rho, _ in vertical_lines])
                    gap_center_x = int(abs(vertical_lines[center_line_idx][0]))
                    return gap_center_x, None, None

            # Fallback
            return edge_image.shape[1] // 2, None, None

    def measure_gap_at_multiple_heights(self, image, gap_center_x, num_measurements=30):
        """Measure gap width at multiple heights using refined method"""
        gray, enhanced, bilateral = self.preprocess_for_vertical_gap(image)
        height, width = gray.shape

        measurements = []
        positions = []
        gap_coordinates = []

        # Define scan region (avoid top and bottom edges)
        start_y = int(height * 0.15)
        end_y = int(height * 0.85)

        # Scan at multiple heights
        for y in np.linspace(start_y, end_y, num_measurements, dtype=int):
            # Extract horizontal line profile centered around gap
            search_width = 100 
            x_start = max(0, gap_center_x - search_width)
            x_end = min(width, gap_center_x + search_width)

            line_profile = bilateral[y, x_start:x_end]

            # Calculate gradient to find edges
            gradient = np.gradient(line_profile.astype(float))

            # Find dark region (gap) - look for valley in intensity
            # Smooth the profile
            smoothed = signal.savgol_filter(line_profile.astype(float),
                                           window_length=min(11, len(line_profile)-1 if len(line_profile) % 2 == 0 else len(line_profile)),
                                           polyorder=2)

            # Find local minimum (darkest point - likely gap center)
            center_idx = len(smoothed) // 2
            search_range = 40
            start_search = max(0, center_idx - search_range)
            end_search = min(len(smoothed), center_idx + search_range)

            local_min_idx = start_search + np.argmin(smoothed[start_search:end_search])

            # Find edges of gap by looking for intensity increase on both sides
            threshold_intensity = smoothed[local_min_idx] + 0.15 * (np.max(smoothed) - np.min(smoothed))

            # Search left edge
            gap_left = local_min_idx
            for i in range(local_min_idx, max(0, local_min_idx - 30), -1):
                if smoothed[i] > threshold_intensity:
                    gap_left = i
                    break

            # Search right edge
            gap_right = local_min_idx
            for i in range(local_min_idx, min(len(smoothed), local_min_idx + 30)):
                if smoothed[i] > threshold_intensity:
                    gap_right = i
                    break

            # Convert back to image coordinates
            gap_left_abs = x_start + gap_left
            gap_right_abs = x_start + gap_right
            gap_width = gap_right - gap_left

            # Only record if gap seems reasonable (not too small or too large)
            if 2 < gap_width < 50:
                measurements.append(gap_width)
                positions.append(y)
                gap_coordinates.append((gap_left_abs, gap_right_abs))

        return measurements, positions, gap_coordinates, gray, enhanced, bilateral

    def pixels_to_mm(self, pixels):
        if self.mm_per_pixel:
            return pixels * self.mm_per_pixel
        elif self.calibration_factor:
            return pixels / self.calibration_factor
        else:
            return pixels

    def analyze_image(self, image_path, visualize=True, actual_gap_mm=None, quick_mode=False):
        """Complete analysis of gap in image
           quick_mode: when True, skip heavy visualization and coordinate map creation."""
        # Load image
        image = cv2.imread(image_path, cv2.IMREAD_COLOR)

        if image is None:
            print(f"Failed to load image: {image_path}")
            return None

        # Downscale large images to target megapixels (12 MP) for efficient measurement and avoid crashes
        image = self.resize_image_to_target(image, target_mp=12.0)

        print(f"\nLoaded image: {image_path}")
        print(f" Image shape: {image.shape} (height, width, channels)")
        if self.calibration_factor:
            print(f" Calibration factor set: {self.calibration_factor} px/mm")
            print(f" mm_per_pixel: {self.mm_per_pixel:.6f} mm/pixel")
        else:
            print(" No calibration factor set. Measurements remain in pixels.")

        print("=" * 60)
        print("WELD GAP ANALYSIS - Improved Detection")
        print("=" * 60)

        # Step 1: Preprocessing
        gray, enhanced, bilateral = self.preprocess_for_vertical_gap(image)
        print("\n Image preprocessing complete")

        # Step 2: Detect vertical edges
        if quick_mode:
            edge_image = self.detect_vertical_edge_fast(bilateral)
        else:
            edge_image = self.detect_vertical_edge(bilateral)
        print(" Vertical edge detection complete")

        # Step 3: Find gap center line
        gap_center_x, projection, smoothed_proj = self.find_gap_center_line(edge_image)
        print(f" Gap center detected at X = {gap_center_x}")  

        # Step 4: Measure gap at multiple heights
        measurements, positions, gap_coordinates, *_ = self.measure_gap_at_multiple_heights(
            image, gap_center_x, num_measurements=30
        )

        print(f" Measurements taken at {len(measurements)} positions")
        print("\n" + "=" * 60)
        print("MEASUREMENT RESULTS")
        print("=" * 60)

        if len(measurements) > 0:
            # Calculate statistics
            mean_width = np.mean(measurements)
            median_width = np.median(measurements)
            std_width = np.std(measurements)
            min_width = np.min(measurements)
            max_width = np.max(measurements)

            print(f"\nStatistics (in pixels):")
            print(f"  Mean gap width:     {mean_width:.2f} px")
            print(f"  Median gap width:   {median_width:.2f} px")
            print(f"  Std deviation:      {std_width:.2f} px")
            print(f"  Min gap width:      {min_width:.2f} px")
            print(f"  Max gap width:      {max_width:.2f} px")

            # Simple calibration conversion
            print("\n" + "=" * 60)
            print("CALIBRATION & MM CONVERSION")
            print("=" * 60)
            print("To convert to millimeters, measure the gap with calipers.")
            print("Then use: mm = pixels × (actual_mm / detected_avg_px)")

            # Check if calibration exists, otherwise use example
            if self.calibration_factor:
                mm_per_pixel = self.mm_per_pixel
                actual_gap_mm = self.pixels_to_mm(mean_width)
                print(f"\n✓ Using calibration factor: {self.calibration_factor:.4f} px/mm")
                print(f"✓ Resolution: {mm_per_pixel:.6f} mm/pixel")
            if actual_gap_mm is not None:
                detected_px = mean_width
                mm_per_pixel = actual_gap_mm / detected_px

                print(f"\n Using USER calibration:")
                print(f"   Detected average:  {detected_px:.2f} px")
                print(f"   Actual gap:        {actual_gap_mm:.3f} mm")
                print(f"   Conversion factor: {mm_per_pixel:.6f} mm/pixel")
            else:
                print("\n No calibration provided.")
                print("Showing pixel results only (no mm conversion).")
                mm_per_pixel = None

                print(f"\n EXAMPLE Calibration (update with your actual measurement):")
                print(f"   Detected average:  {detected_px:.2f} px")
                print(f"   Actual gap (est):  {actual_gap_mm:.3f} mm  ← MEASURE WITH CALIPERS")
                print(f"   Conversion factor: {mm_per_pixel:.6f} mm/pixel")

            # Convert all measurements to mm
            if mm_per_pixel:
                measurements_mm = np.array(measurements) * mm_per_pixel
            else:
                measurements_mm = None

            if measurements_mm is not None:
              print(f"\nStatistics (in millimeters):")
              print(f"  Mean gap width:     {np.mean(measurements_mm):.4f} mm")
              print(f"  Median gap width:   {np.median(measurements_mm):.4f} mm")
              print(f"  Std deviation:      {np.std(measurements_mm):.4f} mm")
              print(f"  Min gap width:      {np.min(measurements_mm):.4f} mm")
              print(f"  Max gap width:      {np.max(measurements_mm):.4f} mm")

            print(f"\n{'Measurement':<12} {'Y Position':<12} {'Gap (px)':<15} {'Gap (mm)':<15}")
            print("-" * 60)
            for i, (width, y_pos, width_mm) in enumerate(zip(measurements, positions, measurements_mm), 1):
                print(f"{i:<12d} {y_pos:<12d} {width:<15.2f} {width_mm:<15.4f}")

            # Visualization
            if visualize and not quick_mode:
                self.create_visualization(image, gray, enhanced, bilateral, edge_image,
                                        gap_center_x, measurements, positions,
                                        gap_coordinates, projection, smoothed_proj)
                self.create_coordinate_reference_map(image, gap_center_x, measurements,
                                                     positions, gap_coordinates)
        else:
            print("\n No valid gap measurements found!")
            print("Possible issues:")
            print("  - Gap may be too narrow or too wide")
            print("  - Image quality may be insufficient")
            print("  - Gap may not be vertical enough")

        return measurements, positions, gap_coordinates

    def create_visualization(self, image, gray, enhanced, bilateral, edge_image,
                           gap_center_x, measurements, positions, gap_coordinates,
                           projection, smoothed_proj):
        """Create comprehensive visualization of gap detection"""

        # Create main figure with multiple subplots
        fig = plt.figure(figsize=(20, 12))
        gs = fig.add_gridspec(3, 3, hspace=0.3, wspace=0.3)

        height, width = image.shape[:2]

        # 1. Original image with scale
        ax1 = fig.add_subplot(gs[0, 0])
        ax1.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
        ax1.set_title('Original Image', fontsize=12, fontweight='bold')
        ax1.set_xlabel('X coordinate (pixels)', fontsize=10)
        ax1.set_ylabel('Y coordinate (pixels)', fontsize=10)
        ax1.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        # Add coordinate labels
        ax1.set_xticks(np.linspace(0, width, 5))
        ax1.set_yticks(np.linspace(0, height, 5))

        # 2. Enhanced image with scale
        ax2 = fig.add_subplot(gs[0, 1])
        ax2.imshow(enhanced, cmap='gray')
        ax2.set_title('Enhanced (CLAHE)', fontsize=12, fontweight='bold')
        ax2.set_xlabel('X coordinate (pixels)', fontsize=10)
        ax2.set_ylabel('Y coordinate (pixels)', fontsize=10)
        ax2.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax2.set_xticks(np.linspace(0, width, 5))
        ax2.set_yticks(np.linspace(0, height, 5))

        # 3. Edge detection with scale
        ax3 = fig.add_subplot(gs[0, 2])
        ax3.imshow(edge_image, cmap='hot')
        ax3.axvline(x=gap_center_x, color='cyan', linestyle='--', linewidth=2,
                   label=f'Gap Center\n(X={gap_center_x})')
        ax3.set_title('Vertical Edge Detection', fontsize=12, fontweight='bold')
        ax3.set_xlabel('X coordinate (pixels)', fontsize=10)
        ax3.set_ylabel('Y coordinate (pixels)', fontsize=10)
        ax3.legend(loc='upper right', fontsize=9)
        ax3.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax3.set_xticks(np.linspace(0, width, 5))
        ax3.set_yticks(np.linspace(0, height, 5))

        # 4. Original with measurements and coordinate scale
        ax4 = fig.add_subplot(gs[1, :])
        vis_image = image.copy()

        for i, (y, width, coords) in enumerate(zip(positions, measurements, gap_coordinates)):
            gap_left, gap_right = coords

            # Draw measurement line
            cv2.line(vis_image, (gap_left, y), (gap_right, y), (0, 255, 0), 2)

            # Draw markers
            cv2.circle(vis_image, (gap_left, y), 4, (255, 0, 0), -1)
            cv2.circle(vis_image, (gap_right, y), 4, (0, 0, 255), -1)

            # Add width text (every 3rd measurement to avoid clutter)
            if i % 3 == 0:
                text_x = gap_right + 10
                if text_x + 80 > image.shape[1]:
                    text_x = gap_left - 90

                cv2.putText(vis_image, f"{width:.1f}px",
                           (text_x, y),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 2)

        # Draw center line
        cv2.line(vis_image, (gap_center_x, 0), (gap_center_x, image.shape[0]),
                (255, 0, 255), 2)

        # Add coordinate labels on the center line
        for y_label in [50, height//2, height-50]:
            cv2.putText(vis_image, f"Y={y_label}",
                       (gap_center_x + 10, y_label),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 255), 2)

        ax4.imshow(cv2.cvtColor(vis_image, cv2.COLOR_BGR2RGB))
        ax4.set_title(f'Gap Measurements (n={len(measurements)}) | Gap Center at X={gap_center_x}',
                     fontsize=14, fontweight='bold')
        ax4.set_xlabel('X coordinate (pixels)', fontsize=11)
        ax4.set_ylabel('Y coordinate (pixels)', fontsize=11)

        # Add major grid
        ax4.grid(True, alpha=0.3, linestyle='-', linewidth=1, color='white')

        # Set coordinate ticks
        x_ticks = np.linspace(0, width, 9)
        y_ticks = np.linspace(0, height, 9)
        ax4.set_xticks(x_ticks)
        ax4.set_yticks(y_ticks)
        ax4.set_xticklabels([f'{int(x)}' for x in x_ticks], fontsize=9)
        ax4.set_yticklabels([f'{int(y)}' for y in y_ticks], fontsize=9)

        # Add minor grid for finer resolution
        ax4.set_xticks(np.linspace(0, width, 17), minor=True)
        ax4.set_yticks(np.linspace(0, height, 17), minor=True)
        ax4.grid(which='minor', alpha=0.15, linestyle=':', linewidth=0.5, color='yellow')

        # 5. Gap width profile
        ax5 = fig.add_subplot(gs[2, 0])
        ax5.plot(measurements, positions, 'bo-', linewidth=2, markersize=6)
        ax5.axvline(x=np.mean(measurements), color='red', linestyle='--',
                   linewidth=2, label=f'Mean: {np.mean(measurements):.2f}px')
        ax5.set_xlabel('Gap Width (pixels)', fontsize=11)
        ax5.set_ylabel('Y Position (pixels)', fontsize=11)
        ax5.set_title('Gap Width Profile', fontsize=12, fontweight='bold')
        ax5.grid(True, alpha=0.3)
        ax5.legend()
        ax5.invert_yaxis()

        # 6. Histogram of gap widths
        ax6 = fig.add_subplot(gs[2, 1])
        ax6.hist(measurements, bins=20, color='steelblue', edgecolor='black', alpha=0.7)
        ax6.axvline(x=np.mean(measurements), color='red', linestyle='--',
                   linewidth=2, label=f'Mean: {np.mean(measurements):.2f}px')
        ax6.axvline(x=np.median(measurements), color='green', linestyle='--',
                   linewidth=2, label=f'Median: {np.median(measurements):.2f}px')
        ax6.set_xlabel('Gap Width (pixels)', fontsize=11)
        ax6.set_ylabel('Frequency', fontsize=11)
        ax6.set_title('Gap Width Distribution', fontsize=12, fontweight='bold')
        ax6.legend()
        ax6.grid(True, alpha=0.3)

        # 7. Vertical projection (if available)
        if projection is not None and smoothed_proj is not None:
            ax7 = fig.add_subplot(gs[2, 2])
            x_coords = np.arange(len(projection))
            ax7.plot(x_coords, projection, 'b-', alpha=0.3, label='Raw')
            ax7.plot(x_coords, smoothed_proj, 'r-', linewidth=2, label='Smoothed')
            ax7.axvline(x=gap_center_x, color='green', linestyle='--',
                       linewidth=2, label='Detected Gap')
            ax7.set_xlabel('X Position (pixels)', fontsize=11)
            ax7.set_ylabel('Edge Strength', fontsize=11)
            ax7.set_title('Vertical Projection Analysis', fontsize=12, fontweight='bold')
            ax7.legend()
            ax7.grid(True, alpha=0.3)

        plt.savefig('outputs/improved_gap_analysis.png', dpi=150, bbox_inches='tight')
        print(f"\n✓ Visualization saved: improved_gap_analysis.png")
        plt.close(fig)

        # Create zoomed view
        if len(gap_coordinates) > 0:
            self.create_zoomed_view(vis_image, gap_coordinates, positions, measurements)

    def create_coordinate_reference_map(self, image, gap_center_x, measurements,
                                       positions, gap_coordinates):
        """Create a detailed coordinate reference map with rulers"""
        fig, ax = plt.subplots(1, 1, figsize=(16, 12))

        height, width = image.shape[:2]

        # Create annotated image
        ref_image = image.copy()

        # Draw all measurements
        for i, (y, width_px, coords) in enumerate(zip(positions, measurements, gap_coordinates)):
            gap_left, gap_right = coords

            # Measurement line
            cv2.line(ref_image, (gap_left, y), (gap_right, y), (0, 255, 0), 2)
            cv2.circle(ref_image, (gap_left, y), 5, (255, 0, 0), -1)
            cv2.circle(ref_image, (gap_right, y), 5, (0, 0, 255), -1)

            # Coordinate annotations (every 5th measurement)
            if i % 5 == 0:
                # Left edge annotation
                cv2.putText(ref_image, f"({gap_left},{y})",
                           (gap_left - 90, y - 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 0, 0), 1)

                # Right edge annotation
                cv2.putText(ref_image, f"({gap_right},{y})",
                           (gap_right + 10, y - 10),
                           cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 255), 1)

        # Draw center line with coordinates
        cv2.line(ref_image, (gap_center_x, 0), (gap_center_x, height),
                (255, 0, 255), 3)

        # Add coordinate markers along center line
        for y_mark in range(0, height, height//10):
            cv2.circle(ref_image, (gap_center_x, y_mark), 6, (255, 255, 0), -1)
            cv2.putText(ref_image, f"X={gap_center_x}",
                       (gap_center_x + 15, y_mark),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 0), 2)

        # Display image
        ax.imshow(cv2.cvtColor(ref_image, cv2.COLOR_BGR2RGB))
        ax.set_title('Coordinate Reference Map with Measurement Points',
                    fontsize=16, fontweight='bold', pad=20)

        # Add detailed grid
        ax.grid(True, which='major', alpha=0.6, linestyle='-', linewidth=1.5, color='white')
        ax.grid(which='minor', alpha=0.3, linestyle=':', linewidth=0.5, color='yellow')

        # X-axis configuration
        x_major_ticks = np.arange(0, width, width//10 if width//10 > 0 else 50)
        x_minor_ticks = np.arange(0, width, width//20 if width//20 > 0 else 25)
        ax.set_xticks(x_major_ticks)
        ax.set_xticks(x_minor_ticks, minor=True)
        ax.set_xlabel('X Coordinate (pixels)', fontsize=14, fontweight='bold', labelpad=10)

        # Y-axis configuration
        y_major_ticks = np.arange(0, height, height//10 if height//10 > 0 else 50)
        y_minor_ticks = np.arange(0, height, height//20 if height//20 > 0 else 25)
        ax.set_yticks(y_major_ticks)
        ax.set_yticks(y_minor_ticks, minor=True)
        ax.set_ylabel('Y Coordinate (pixels)', fontsize=14, fontweight='bold', labelpad=10)

        # Rotate x-axis labels for better readability
        ax.tick_params(axis='x', rotation=45, labelsize=10)
        ax.tick_params(axis='y', labelsize=10)

        # Add legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='green', label='Measurement Line'),
            Patch(facecolor='blue', label='Left Edge (Start)'),
            Patch(facecolor='red', label='Right Edge (End)'),
            Patch(facecolor='magenta', label=f'Gap Center (X={gap_center_x})'),
            Patch(facecolor='yellow', label='Coordinate Markers')
        ]
        ax.legend(handles=legend_elements, loc='upper right', fontsize=10,
                 framealpha=0.9, edgecolor='black', fancybox=True)

        # Add info box
        info_text = (
            f"IMAGE DIMENSIONS: {width} × {height} pixels\n"
            f"GAP CENTER LINE: X = {gap_center_x} px\n"
            f"MEASUREMENTS: {len(measurements)} points\n"
            f"Y RANGE: {min(positions)} to {max(positions)} px"
        )
        ax.text(0.02, 0.98, info_text, transform=ax.transAxes,
               fontsize=11, verticalalignment='top',
               bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.9,
                        edgecolor='black', linewidth=2),
               family='monospace')

        plt.tight_layout()
        plt.savefig('outputs/coordinate_reference_map.png', dpi=200, bbox_inches='tight')
        print(f"✓ Coordinate reference map saved: coordinate_reference_map.png")
        plt.close(fig)
    def save_visualization(self, output_path):
      if hasattr(self, "visualized_image"):
        cv2.imwrite(output_path, self.visualized_image)

    def create_zoomed_view(self, vis_image, gap_coordinates, positions, measurements):
        """Create a detailed zoomed view of the gap region"""
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 14))

        # Determine crop region
        all_lefts = [coord[0] for coord in gap_coordinates]
        all_rights = [coord[1] for coord in gap_coordinates]

        gap_center_x = (min(all_lefts) + max(all_rights)) // 2
        crop_width = 150  # Show 150 pixels on each side

        x_start = max(0, gap_center_x - crop_width)
        x_end = min(vis_image.shape[1], gap_center_x + crop_width)
        y_start = max(0, min(positions) - 30)
        y_end = min(vis_image.shape[0], max(positions) + 30)

        cropped = vis_image[y_start:y_end, x_start:x_end]

        # Left panel - zoomed view with grid
        ax1.imshow(cv2.cvtColor(cropped, cv2.COLOR_BGR2RGB))
        ax1.set_title('Zoomed View - Gap Detail', fontsize=14, fontweight='bold')
        ax1.set_xlabel(f'X coordinate (pixels, origin at {x_start})', fontsize=10)
        ax1.set_ylabel(f'Y coordinate (pixels, origin at {y_start})', fontsize=10)

        # Add grid and coordinate labels
        ax1.grid(True, alpha=0.4, linestyle='-', linewidth=1, color='white')
        ax1.grid(which='minor', alpha=0.2, linestyle=':', linewidth=0.5, color='yellow')

        # Set ticks to show actual image coordinates
        crop_height, crop_width_px = cropped.shape[:2]
        x_ticks_local = np.linspace(0, crop_width_px, 7)
        y_ticks_local = np.linspace(0, crop_height, 7)

        # Convert local coordinates to global image coordinates
        x_labels = [f'{int(x_start + x)}' for x in x_ticks_local]
        y_labels = [f'{int(y_start + y)}' for y in y_ticks_local]

        ax1.set_xticks(x_ticks_local)
        ax1.set_yticks(y_ticks_local)
        ax1.set_xticklabels(x_labels, fontsize=9)
        ax1.set_yticklabels(y_labels, fontsize=9)

        # Add minor ticks
        ax1.set_xticks(np.linspace(0, crop_width_px, 13), minor=True)
        ax1.set_yticks(np.linspace(0, crop_height, 13), minor=True)

        # Right panel - measurement details table
        ax2.axis('off')
        ax2.set_title('Measurement Details', fontsize=14, fontweight='bold')

        # Create detailed measurement table
        table_data = []

        table_data.append(['#', 'Y Pos', 'Left X', 'Right X', 'Width', 'Status'])
        table_data.append(['', '(px)', '(px)', '(px)', '(px)', ''])

        for i, (y, width, coords) in enumerate(zip(positions, measurements, gap_coordinates), 1):
            gap_left, gap_right = coords

            # Convert to mm (IMPORTANT)
            mm = width * self.mm_per_pixel if hasattr(self, 'mm_per_pixel') and self.mm_per_pixel else width

            # Status logic
            status = 'OK' if mm <= 0.5 else 'NG'

            # ALWAYS 6 elements
            table_data.append([
                str(i),
                str(int(y)),
                str(int(gap_left)),
                str(int(gap_right)),
                f"{mm:.3f}",
                status
            ])

        # Create table
        table = ax2.table(
            cellText=table_data,
            cellLoc='center',
            loc='center',
            bbox=[0.1, 0.1, 0.8, 0.8]
        )

        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 2)

        # Style header rows
        for i in range(2):
            for j in range(6):  # updated to 6 columns
                cell = table[(i, j)]
                cell.set_facecolor('#4472C4')
                cell.set_text_props(weight='bold', color='white')

        # Alternate row colors
        for i in range(2, len(table_data)):
            for j in range(6):  # updated to 6 columns
                cell = table[(i, j)]
                if i % 2 == 0:
                    cell.set_facecolor('#E7E6E6')
                else:
                    cell.set_facecolor('#F2F2F2')
        # Add statistics box
        stats_text = f"""
        STATISTICS
        ─────────────────
        Mean:      {np.mean(measurements):.2f} px
        Median:    {np.median(measurements):.2f} px
        Std Dev:   {np.std(measurements):.2f} px
        Min:       {np.min(measurements):.2f} px
        Max:       {np.max(measurements):.2f} px

        Gap Center: X = {gap_center_x} px
        Measurements: {len(measurements)}
        """

        ax2.text(0.5, 0.05, stats_text, transform=ax2.transAxes,
                fontsize=10, verticalalignment='bottom', horizontalalignment='center',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
                family='monospace')

        plt.tight_layout()
        plt.savefig('outputs/gap_zoomed_detail.png', dpi=150, bbox_inches='tight')
        print(f"✓ Zoomed view saved: gap_zoomed_detail.png")
        plt.close(fig)

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog
from tkinter import ttk
from PIL import Image, ImageTk
import os
import shutil
import io
import sys

class GapApp:

    def __init__(self, root):

        self.root = root
        self.root.title("Gap Measurement System")
        self.root.configure(bg="#f0f4f8")
        self.root.geometry("1600x950")

        self.detector = ImprovedGapDetector()

        self.image_path = None

        title = tk.Label(root, text="Weld Gap Measurement System", font=("Arial", 20, "bold"), bg="#f0f4f8", fg="#2a2f45")
        title.pack(pady=12)

        btn_frame = tk.Frame(root, bg="#FFFFFF")
        btn_frame.pack(pady=8)

        btn_style = {"width":18, "bg":"#04418B", "fg":"white", "font":("Arial", 11, "bold"), "bd":0, "relief":"flat", "activebackground":"#1f5ab3", "activeforeground":"white", "highlightthickness":0}

        def make_button(text, cmd, col):
            button = tk.Button(btn_frame, text=text, command=cmd, **btn_style)
            button.grid(row=0, column=col, padx=8, pady=2)
            button.bind("<Enter>", lambda e: button.configure(bg="#1f5ab3"))
            button.bind("<Leave>", lambda e: button.configure(bg="#2a76d2"))
            return button

        make_button("Load Image", self.load_image, 0)
        make_button("Capture Photo", self.capture_photo, 1)
        make_button("Run Analysis", self.run_analysis, 2)
        make_button("Save Report", self.save_report, 3)

        # Add calibration input
        cal_frame = tk.Frame(root)
        cal_frame.pack(pady=5)
        tk.Label(cal_frame, text="Actual Gap (mm):", font=("Arial",12)).grid(row=0, column=0, padx=5)
        self.actual_mm_entry = tk.Entry(cal_frame, width=10, font=("Arial",12))
        self.actual_mm_entry.grid(row=0, column=1, padx=5)
        self.actual_mm_entry.insert(0, "0.0")  # default value

        # Add speed-up mode toggle (skip heavy plotting)
        self.fast_mode_var = tk.BooleanVar(value=True)
        self.fast_mode_check = tk.Checkbutton(cal_frame, text="Fast mode (approximate)", variable=self.fast_mode_var, bg="#f0f4f8", font=("Arial", 10, "bold"))
        self.fast_mode_check.grid(row=0, column=2, padx=10)

        # Main content frame with grid layout
        content_frame = tk.Frame(root)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        content_frame.grid_rowconfigure(0, weight=1)
        content_frame.grid_rowconfigure(1, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=1)

        # Left top: input image
        self.preview = tk.Label(content_frame)
        self.preview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

        # Right: coordinate reference map image
        self.img3 = tk.Label(content_frame)
        self.img3.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=5, pady=5)
        
        # Left bottom: output table
        self.tree = ttk.Treeview(
            content_frame,
            columns=("num", "y", "px", "mm", "status"),
            show="headings",
            height=15
        )

        self.tree.heading("num", text="Measurement")
        self.tree.heading("y", text="Y Position")
        self.tree.heading("px", text="Gap Width (px)")
        self.tree.heading("mm", text="Gap Width (mm)")
        self.tree.heading("status", text="Status")   # NEW COLUMN

        self.tree.column("num", width=80)
        self.tree.column("y", width=80)
        self.tree.column("px", width=80)
        self.tree.column("mm", width=80)
        self.tree.column("status", width=80)  # NEW COLUMN

        self.tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)

                # Status label above content
        self.status_label = tk.Label(root, text="", font=("Arial", 12, "bold"), bg="#f0f4f8", fg="#155fa0")
        self.status_label.pack(pady=8)

    def load_image(self):

        path = filedialog.askopenfilename(
            filetypes=[("Images","*.png *.jpg *.jpeg *.bmp")]
        )

        if not path:
            return

        self.image_path = path

        img = Image.open(path)
        img = img.resize((300,300))

        tkimg = ImageTk.PhotoImage(img)

        self.preview.configure(image=tkimg)
        self.preview.image = tkimg

        # Restore the preview to grid layout (in case it was hidden after previous analysis)
        self.preview.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        
        # Clear previous analysis results
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.status_label.config(text="")

    def capture_photo(self):
        """Capture photo from camera"""
        cap = cv2.VideoCapture(1)
        
        if not cap.isOpened():
            messagebox.showerror("Error", "Could not access camera. Make sure camera is connected.")
            return
        
        # Create a window for camera preview
        while True:
            ret, frame = cap.read()
            if not ret:
                messagebox.showerror("Error", "Failed to capture image from camera")
                break

            # Display the frame
            cv2.imshow('Camera - Press SPACE to capture, ESC to cancel', frame)
            
            key = cv2.waitKey(1) & 0xFF
            if key == 32:  # SPACE key
                # Save captured image
                temp_path = "temp_camera_capture.png"
                cv2.imwrite(temp_path, frame)
                self.image_path = temp_path
                
                # Convert BGR to RGB for PIL
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img = Image.fromarray(frame_rgb)
                img = img.resize((300,300))
                
                tkimg = ImageTk.PhotoImage(img)
                self.preview.configure(image=tkimg)
                self.preview.image = tkimg
                
                messagebox.showinfo("Success", "Photo captured successfully")
                break
            elif key == 27:  # ESC key
                messagebox.showinfo("Cancelled", "Camera capture cancelled")
                break
    
        cap.release()
        cv2.destroyAllWindows()

    def run_analysis(self):

        if not self.image_path:
            messagebox.showerror("Error","Load image first")
            return

        if not os.path.exists("outputs"):
            os.makedirs("outputs")

        actual_mm_str = self.actual_mm_entry.get().strip()
        actual_gap_mm = None
        if actual_mm_str:
            try:
                actual_gap_mm = float(actual_mm_str)
                if actual_gap_mm <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Invalid actual gap value.")
                return

        self.status_label.config(text="Analyzing... Please wait.")
        self.root.update()

        old_stdout = sys.stdout
        sys.stdout = captured_output = io.StringIO()

        fast_mode = self.fast_mode_var.get()

        try:
            results = self.detector.analyze_image(
                self.image_path,
                visualize=not fast_mode,
                actual_gap_mm=actual_gap_mm,
                quick_mode=fast_mode
            )
        finally:
            sys.stdout = old_stdout

        if results and len(results) >= 3:
            measurements, positions, gap_coordinates = results[:3]

            # Clear table
            for item in self.tree.get_children():
                self.tree.delete(item)

            mm_per_pixel = None
            if actual_gap_mm and len(measurements) > 0:
                mean_px = np.mean(measurements)
                if mean_px > 0:
                    mm_per_pixel = actual_gap_mm / mean_px

            has_over_limit = False

            # Configure tags
            self.tree.tag_configure('ok', background='#90ee90')         # green
            self.tree.tag_configure('over_limit', background='#ff6b6b', foreground='white')  # red

            for i, (px, y) in enumerate(zip(measurements, positions), 1):

                mm = px * mm_per_pixel if mm_per_pixel else 0.0
                mm_str = f"{mm:.4f}" if mm_per_pixel else "N/A"

                # Status logic
                if mm_per_pixel:
                    status = "OK" if mm <= 0.5 else "NG"
                else:
                    status = "N/A"

                if mm_per_pixel and mm > 0.5:
                    has_over_limit = True

                # Insert WITH status
                row_id = self.tree.insert(
                    "",
                    "end",
                    values=(i, int(y), f"{px:.2f}", mm_str, status)
                )

                # Apply row color
                if status == "NG":
                    self.tree.item(row_id, tags=('over_limit',))
                else:
                    self.tree.item(row_id, tags=('ok',))

            # Overall result
            if has_over_limit:
                self.status_label.config(
                    text="❌ FAIL - Gap exceeds 0.5mm",
                    fg='red'
                )
            else:
                self.status_label.config(
                    text="✅ PASS - All gaps within limit",
                    fg='green'
                )

        else:
            self.status_label.config(text="Analysis failed.")

        # Layout adjustments
        self.preview.grid_forget()
        self.tree.grid(row=0, column=0, rowspan=2, sticky="nsew")

        if not fast_mode:
            self.show_outputs()
        else:
            self.img3.configure(image='')
            self.img3.image = None

    def show_outputs(self):

        # Show only coordinate reference map
        file = "outputs/coordinate_reference_map.png"

        if os.path.exists(file):

            img = Image.open(file)
            img = img.resize((600,700))

            tkimg = ImageTk.PhotoImage(img)

            self.img3.configure(image=tkimg)
            self.img3.image = tkimg


    def save_report(self):

        from tkinter import filedialog, messagebox, simpledialog
        parent_folder = filedialog.askdirectory(title="Select parent folder to save report")
        if not parent_folder:
            return

        # Ask for new folder name
        new_folder_name = simpledialog.askstring("New Folder", "Enter folder name for the report:")

        if not new_folder_name:
            return

        # Create new folder
        report_folder = os.path.join(parent_folder, new_folder_name)
        try:
            os.makedirs(report_folder, exist_ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create folder: {str(e)}")
            return

        # Copy output images
        files = [
            "outputs/gap_zoomed_detail.png",
            "outputs/coordinate_reference_map.png",
            "outputs/improved_gap_analysis.png"
        ]

        for file in files:
            if os.path.exists(file):
                shutil.copy(file, report_folder)

        # Extract Treeview data
        rows = []
        for item_id in self.tree.get_children():
            row = self.tree.item(item_id)['values']
            rows.append(row)

        if rows:
            headers = [self.tree.heading(c)['text'] for c in self.tree['columns']]
            excel_path = os.path.join(report_folder, 'gap_measurements.xlsx')
            csv_path = os.path.join(report_folder, 'gap_measurements.csv')

            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font

                wb = Workbook()
                ws = wb.active
                ws.title = 'Gap Measurements'

                # Styles
                left_align = Alignment(horizontal='left')
                bold_font = Font(bold=True)

                # Add headers
                ws.append(headers)

                # Style header row
                for cell in ws[1]:
                    cell.alignment = left_align
                    cell.font = bold_font

                # Add data rows
                for row in rows:
                    ws.append(row)

                # Apply left alignment to all cells
                for row_cells in ws.iter_rows():
                    for cell in row_cells:
                        cell.alignment = left_align

                # Auto-adjust column width
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                # Save Excel file
                wb.save(excel_path)

                messagebox.showinfo(
                    'Saved',
                    f"Report saved to {report_folder}\nExcel: {excel_path}"
                )

            except ImportError:
                # Fallback to CSV
                import csv
                with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(headers)
                    writer.writerows(rows)

                messagebox.showinfo(
                    'Saved',
                    f"Report saved to {report_folder}\nCSV: {csv_path}"
                )

            except Exception as e:
                messagebox.showerror('Error', f"Failed to save table: {e}")

        else:
            messagebox.showinfo(
                'Saved',
                f"Report images saved to {report_folder}\nNo table rows to store."
            )
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Gap Measurement System")

    # Open in maximized window (preferred)
    try:
        root.state("zoomed")
    except Exception:
        # fallback to manual max size
        root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")

    app = GapApp(root)

    # Handle window close properly
    root.protocol("WM_DELETE_WINDOW", root.quit)

    root.mainloop()